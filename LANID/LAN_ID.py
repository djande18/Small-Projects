import os
import openpyxl



workbook = openpyxl.Workbook()

sheet = workbook.active

LANID_header = sheet.cell(row=1,column=1)
LANID_header.value = ("LANID")

Name_header = sheet.cell(row=1,column=2)
Name_header.value = ("Name")

# This program accepts text files that contain a list of people assigned to a certain drive. It check all files contained in the same folder that this script is in. 
# The output will be a list of LANIDs and a list of names that can be pasted in an excel file

def LAN_Name(file):
    
    #Open text file
    with open(file, 'r') as file:
        
        #Create empty lists for collecting the desired info
        LANIDs = []
        names = []
        

        for i,line in enumerate(file):
            #Ignore the first five lines to get to the names/LANIDs
            if i < 5:
                continue
            else:
                #Separate the LANID from the name using an index
                space_index = line.find(" ")

                LANID = line[:space_index]

                #Separate the name from the rest of the line using an index
                mfad_index = line.find("mfad")

                name = line[space_index+1:mfad_index]

            LANIDs.append(LANID)
            cell1 = sheet.cell(row=i-3,column=1)
            cell1.value = f"{LANID}"

            names.append(name)
            cell2 = sheet.cell(row=i-3,column=2)
            cell2.value = f"{name}"

    #Perform a minor verification that the length of the name list is the same as the length of the LANID list
    if len(LANIDs) == len(names):
        print("Confirmed")
    else:
        print("Doesn't match up")

    print(f"*******\nLANIDs \n_________\nCount:\n {len(LANIDs)}\n")

    for i in LANIDs:
        print(i)

    print(f"\nNames \n_________\nCount:\n {len(names)}\n")

    for i in names:
        print(i)

current_directory = os.path.dirname(os.path.abspath(__file__))



file_names = os.listdir(current_directory)

#Create a spreadsheet for each of the .txt files
for file_name in file_names:
    
    index = file_name.find(".")
    if file_name[index:] == ".txt":
        
        #only create a workbook if it's a .txt file
        workbook_filename = file_name
        print(f"\n{file_name}")
        LAN_Name(file_name)


        workbook.save(f'{workbook_filename} Access.xlsx')